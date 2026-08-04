import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from config import (
    DIR_BAVLI, DIR_YERUSHALMI,VERB_TAGS, NOUN_TAGS, PREPOSITION_TAGS, CONJUNCTION_TAGS,
    EMPHATIC_STATE_TAGS, ABSOLUTE_STATE_TAGS, PLURAL_TAGS, SINGULAR_TAGS, PASSIVE_VERB_TAGS
)

# Parse the encoded URL description into tractate, page, side, and line fields.
def parse_url_location(url_string):

    # Define a regular expression that captures the four location components.
    pattern = r'Masekhet: (\d+), Page: (\w+), Side: (\w+), Line: (\d+)' 
    match = re.search(pattern, str(url_string))
    if match:
        return pd.Series([match.group(1), match.group(2), match.group(3), match.group(4)])

    # Return default values when the expected location format is not found.
    return pd.Series(['0', '0', '0', '0']) 

def load_data():
    """
    Scan the Bavli and Yerushalmi data directories, identify tractates available
    in both collections, load their CSV files, and combine them into one dataset.
    """

    # Retrieve all CSV files currently available in the configured directories.
    _files_bavli = [f for f in os.listdir(DIR_BAVLI) if f.endswith('.csv')] if os.path.exists(DIR_BAVLI) else []
    _files_yerushalmi = [f for f in os.listdir(DIR_YERUSHALMI) if f.endswith('.csv')] if os.path.exists(DIR_YERUSHALMI) else []

    # Normalize filenames and extract the tractate code used to match both collections.
    def _get_bavli_code(filename):
        return filename.replace('df_', '').replace('_csv.csv', '')

    def _get_yer_code(filename):
        return filename.replace('df_yer_', '').replace('_csv.csv', '')

    # Map each tractate code to its source file and identify the shared codes.
    _bavli_map = {_get_bavli_code(f): f for f in _files_bavli}
    _yer_map = {_get_yer_code(f): f for f in _files_yerushalmi}

    _shared_codes = set(_bavli_map.keys()) & set(_yer_map.keys())

    # Build the final lists of corresponding Bavli and Yerushalmi files.
    bavli_tractates_dynamic = [_bavli_map[code] for code in _shared_codes]
    yerushalmi_tractates_dynamic = [_yer_map[code] for code in _shared_codes]
    
    # Load the Bavli files whose tractates also appear in the Yerushalmi dataset.
    list_bavli = []
    print(f"Loading {len(bavli_tractates_dynamic)} shared Bavli tractates automatically...")
    for filename in bavli_tractates_dynamic:
        full_path = os.path.join(DIR_BAVLI, filename)
        if os.path.exists(full_path):
            list_bavli.append(pd.read_csv(full_path))
            
    df_b = pd.concat(list_bavli, ignore_index=True) if list_bavli else pd.DataFrame()
    
    # Load the Yerushalmi files whose tractates also appear in the Bavli dataset.
    list_yer = []
    print(f"Loading {len(yerushalmi_tractates_dynamic)} shared Yerushalmi tractates automatically...")
    for filename in yerushalmi_tractates_dynamic:
        full_path = os.path.join(DIR_YERUSHALMI, filename)
        if os.path.exists(full_path):
            list_yer.append(pd.read_csv(full_path))
            
    df_y = pd.concat(list_yer, ignore_index=True) if list_yer else pd.DataFrame()

    # Assign the dialect classification label to each dataset.
    if not df_b.empty: df_b['target'] = 'Bavli'
    if not df_y.empty: df_y['target'] = 'Yerushalmi'
    
    # Combine the Bavli and Yerushalmi records into a single dataset.
    df = pd.concat([df_b, df_y], ignore_index=True)

    # Extract the encoded location information into four separate columns.
    df[['masekhet', 'page', 'side', 'line']] = df['url'].apply(parse_url_location)

    return df


def extract_features(group):
    """
    Calculate the linguistic feature values for a single textual line based on
    its lexical tags and word-level information.
    """

    # Combine the lexical annotations into one normalized text sequence.
    lex_text = " ".join(group['merged_lexicon'].fillna('').astype(str).tolist()).lower()
    
    # Split the annotation sequence into individual tags so short tags are matched accurately.
    lex_tokens = lex_text.split()
    
    # Use the number of records in the group as the word count for the current line.
    word_count = len(group)

    # Return an empty result when the group contains no words.
    if word_count == 0: return pd.Series()
    
    # Hypothesis 6: Count plural and singular grammatical tags.
    plural_count = sum(1 for t in lex_tokens if t in PLURAL_TAGS)
    singular_count = sum(1 for t in lex_tokens if t in SINGULAR_TAGS)
    total_numbers = plural_count + singular_count

    # Hypothesis 8: Locate all verb tags within the lexical annotation sequence.
    verb_indices = [i for i, t in enumerate(lex_tokens) if t in VERB_TAGS]
    verb_count_total = len(verb_indices)
    
    v_then_noun_count = 0
    v_then_prep_count = 0
    
    # Examine the tag immediately following each identified verb.
    if verb_count_total > 0:
        for i in verb_indices:

            # Ensure that a following tag exists before accessing it.
            if i + 1 < len(lex_tokens):
                next_tag = lex_tokens[i+1]
                if next_tag in NOUN_TAGS:
                    v_then_noun_count += 1
                elif next_tag in PREPOSITION_TAGS:
                    v_then_prep_count += 1
        
        # Calculate transition ratios relative to the total number of verbs.
        v_n_ratio = round(v_then_noun_count / verb_count_total, 4)
        v_p_ratio = round(v_then_prep_count / verb_count_total, 4)
    else:
        v_n_ratio = 0
        v_p_ratio = 0

    features = {

        # Hypothesis 1: Relative frequency of emphatic and absolute nominal states.
        'emphatic_ratio': round(sum(1 for t in lex_tokens if t in EMPHATIC_STATE_TAGS) / word_count, 4),
        'absolute_ratio': round(sum(1 for t in lex_tokens if t in ABSOLUTE_STATE_TAGS) / word_count, 4),
        
        # Hypothesis 2: Relative frequency of prepositions and conjunctions.
        'function_words_ratio': round(sum(1 for t in lex_tokens if t in PREPOSITION_TAGS or t in CONJUNCTION_TAGS) / word_count, 4),
        
        # Hypothesis 3: Ratio of distinct lemmas to the total number of words.
        # Lines containing three words or fewer receive a neutral value of 0.5
        # to reduce the effect of unstable diversity measurements on short lines.
        'lexical_diversity': round(group['Lema'].nunique() / word_count, 4) if word_count > 3 else 0.5,
        
        # Hypothesis 4: Relative frequency of verbs.
        'verb_ratio': round(sum(1 for t in lex_tokens if t in VERB_TAGS) / word_count, 4),
        
        # Hypothesis 5: Relative frequency of passive verb forms.
        'passive_voice_ratio': round(sum(1 for t in lex_tokens if t in PASSIVE_VERB_TAGS) / word_count, 4),        

        # Hypothesis 6: Proportion of plural forms among all identified number forms.
        'plural_ratio': round(plural_count / total_numbers, 4) if total_numbers > 0 else 0.0,

        # Hypothesis 7: Number of words and average word length in the textual line.
        'line_length': word_count,
        'avg_word_len': round(group['text_transformed'].astype(str).apply(len).mean(), 4),

        # Hypothesis 8: Ratios of verb-to-noun and verb-to-preposition tag transitions.
        'v_then_noun_ratio': v_n_ratio,
        'v_then_prep_ratio': v_p_ratio
    }

    return pd.Series(features)


def save_styled_excel(df, output_path):
    """Save the extracted feature table as a formatted Excel workbook."""

    print("Creates a formatted Excel file with borders and stripes...")
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='ResearchData')
    
    workbook = writer.book
    worksheet = workbook['ResearchData']
    
    # Define the visual styles applied to the worksheet.
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(left=Side(style='thin', color='BDC3C7'), 
                         right=Side(style='thin', color='BDC3C7'), 
                         top=Side(style='thin', color='BDC3C7'), 
                         bottom=Side(style='thin', color='BDC3C7'))
    zebra_fill = PatternFill(start_color='F2F4F4', end_color='F2F4F4', fill_type='solid')
    
    # Apply column widths, borders, alignment, header formatting, and alternating row shading.
    for col_num, column_cells in enumerate(worksheet.columns, 1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = 20
        for i, cell in enumerate(column_cells):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if i == 0:
                cell.fill = header_fill
                cell.font = header_font
            elif i % 2 == 0:
                cell.fill = zebra_fill
    writer.close()

if __name__ == "__main__":
    raw_df = load_data()
    print("Research hypothesis analyzer...")

    # Group individual words by textual location and dialect before extracting features.
    group_cols = ['masekhet', 'page', 'side', 'line', 'target']
    final_table = raw_df.groupby(group_cols, group_keys=False).apply(extract_features).reset_index()
    
    # Define the output paths for the formatted and machine-readable feature tables.
    excel_output = os.path.join('Data', 'ready_for_classifier.xlsx')
    csv_output = os.path.join('Data', 'ready_for_classifier.csv')


    save_styled_excel(final_table, excel_output)
    final_table.to_csv(csv_output, index=False, encoding='utf-8-sig')
    
    print(f"\nSuccess! Two files were created in the 'Data' folder:")
    print(f"   -The formatted Excel: {excel_output}")
    print(f"   -CSV file for the model: {csv_output}")